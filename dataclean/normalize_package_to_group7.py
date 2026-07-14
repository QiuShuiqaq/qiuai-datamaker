from __future__ import annotations

import argparse
import json
import shutil
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from json import JSONDecoder
from pathlib import Path

from openpyxl import Workbook, load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from trajectory_scripts.quality_check import validate_session


DELIVERY_XLSX_NAME = "\u8d28\u68c0\u63d0\u4ea4\u8bb0\u5f55_v2.xlsx"
LEGACY_XLSX_NAME = "\u8d28\u68c0\u63d0\u4ea4\u8bb0\u5f55.xlsx"
DELIVERY_SHEET_NAME = "\u8d28\u68c0\u63d0\u4ea4\u8bb0\u5f55"
HEADER_GROUP = "\u5206\u7ec4"
HEADER_SUBMITTER = "\u63d0\u4ea4\u8005"
HEADER_SESSION_ID = "sessionId"
HEADER_SCENE = "\u5bf9\u8bdd\u573a\u666f"
DELIVERY_HEADERS = [HEADER_GROUP, HEADER_SUBMITTER, HEADER_SESSION_ID, HEADER_SCENE]
DEFAULT_SUBMITTER = "Claude\u672c\u673a\u81ea\u8dd1"
DEFAULT_OUTPUT_SUFFIX = "\u7ec47\u6807\u51c6"
REPORT_SUFFIX = "\u5904\u7406\u62a5\u544a"
PLACEHOLDER_SIGNATURE = "Eo8E_placeholder_sig"
LABEL_MODEL = "deepseek/deepseek-v4-pro"


@dataclass
class CandidateResult:
    path: Path
    ok: bool
    errors: list[str]
    warnings: list[str]
    turns: int


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Normalize an old result package into the approved Group7 delivery format."
    )
    parser.add_argument("--source", required=True, help="Source package directory.")
    parser.add_argument(
        "--output",
        help="Output directory. Defaults to a sibling folder named '<source>_组7标准'.",
    )
    parser.add_argument(
        "--submitter",
        default=DEFAULT_SUBMITTER,
        help="Fallback submitter value when the source workbook has no per-session submitter.",
    )
    return parser.parse_args()


def read_submission_metadata(source_root: Path) -> dict[str, dict[str, str]]:
    xlsx_candidates = [
        source_root / DELIVERY_XLSX_NAME,
        source_root / LEGACY_XLSX_NAME,
    ]
    xlsx_path = next((path for path in xlsx_candidates if path.exists()), None)
    if xlsx_path is None:
        return {}

    workbook = load_workbook(xlsx_path, read_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        next(rows, None)
        mapping: dict[str, dict[str, str]] = {}
        for row in rows:
            if not row or len(row) < 4:
                continue
            session_id = str(row[2] or "").strip()
            if not session_id:
                continue
            if session_id in mapping:
                continue
            mapping[session_id] = {
                "submitter": str(row[1] or "").strip(),
                "scene_name": str(row[3] or "").strip(),
            }
        return mapping
    finally:
        workbook.close()


def iter_agent_roots(source_root: Path) -> list[tuple[str, Path]]:
    agents: list[tuple[str, Path]] = []
    for agent_name in ("openclaw", "Hermes"):
        agent_root = source_root / agent_name
        if agent_root.exists() and agent_root.is_dir():
            agents.append((agent_name, agent_root))
    return agents


def build_candidate_result(path: Path) -> CandidateResult:
    ok, errors, warnings, stats = validate_session(path)
    return CandidateResult(
        path=path,
        ok=ok,
        errors=errors,
        warnings=warnings,
        turns=int(stats.get("assistant_turns", 0) or 0),
    )


def choose_best_candidate(paths: list[Path]) -> CandidateResult:
    results = [build_candidate_result(path) for path in paths]

    def sort_key(item: CandidateResult) -> tuple[int, int, int, int, int]:
        exact_name = 1 if "__" not in item.path.name else 0
        return (
            1 if item.ok else 0,
            -len(item.errors),
            -len(item.warnings),
            item.turns,
            exact_name,
        )

    return max(results, key=sort_key)


def decode_first_json_object(text: str) -> dict:
    stripped = text.strip()
    if not stripped:
        return {}
    try:
        parsed = json.loads(stripped)
        return parsed if isinstance(parsed, dict) else {}
    except json.JSONDecodeError:
        decoder = JSONDecoder()
        parsed, _ = decoder.raw_decode(stripped)
        return parsed if isinstance(parsed, dict) else {}


def normalize_thinking_blocks(blocks) -> bool:
    changed = False
    if not isinstance(blocks, list):
        return changed
    for block in blocks:
        if not isinstance(block, dict):
            continue
        if block.get("type") == "thinking" and not str(block.get("signature", "")).strip():
            block["signature"] = PLACEHOLDER_SIGNATURE
            changed = True
    return changed


def normalize_session_json_files(session_dir: Path, agent_name: str) -> str:
    model_name = ""
    for json_file in sorted(session_dir.glob("*.json")):
        if json_file.name == "task_difficulty_justification.json":
            continue

        with json_file.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)

        changed = False
        if not model_name:
            model_name = str(payload.get("request", {}).get("model", "")).strip()

        if agent_name == "openclaw" and "is_garbled" in payload:
            payload.pop("is_garbled", None)
            changed = True

        if agent_name == "Hermes":
            is_garbled = payload.get("is_garbled")
            if not isinstance(is_garbled, bool):
                payload["is_garbled"] = bool(is_garbled)
                changed = True

        if normalize_thinking_blocks(payload.get("response", {}).get("response_data", {}).get("content")):
            changed = True

        request_messages = payload.get("request", {}).get("messages", [])
        if isinstance(request_messages, list):
            for message in request_messages:
                if isinstance(message, dict) and message.get("role") == "assistant":
                    if normalize_thinking_blocks(message.get("content")):
                        changed = True

        if changed:
            with json_file.open("w", encoding="utf-8") as handle:
                json.dump(payload, handle, ensure_ascii=False, indent=2)

    return model_name


def normalize_label_file(session_dir: Path) -> None:
    json_path = session_dir / "task_difficulty_justification.json"
    jsonl_path = session_dir / "task_difficulty_justification.jsonl"

    if json_path.exists():
        with json_path.open("r", encoding="utf-8") as handle:
            source = json.load(handle)
    elif jsonl_path.exists():
        source = decode_first_json_object(jsonl_path.read_text(encoding="utf-8"))
    else:
        source = {}

    timestamp_source = json_path if json_path.exists() else jsonl_path if jsonl_path.exists() else None
    evaluation_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if timestamp_source is not None:
        evaluation_time = datetime.fromtimestamp(timestamp_source.stat().st_mtime).strftime(
            "%Y-%m-%d %H:%M:%S"
        )

    normalized = {
        "session_id": session_dir.name,
        "task_difficulty": str(source.get("task_difficulty") or source.get("difficulty") or "").strip(),
        "justification": str(source.get("justification") or "").strip(),
        "model": str(source.get("model") or LABEL_MODEL).strip() or LABEL_MODEL,
        "evaluation_time": str(source.get("evaluation_time") or evaluation_time).strip() or evaluation_time,
    }

    with json_path.open("w", encoding="utf-8") as handle:
        json.dump(normalized, handle, ensure_ascii=False, indent=2)

    if jsonl_path.exists():
        jsonl_path.unlink()


def write_delivery_xlsx(path: Path, rows: list[dict]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = DELIVERY_SHEET_NAME
    sheet.append(DELIVERY_HEADERS)
    for row in rows:
        sheet.append([row.get(header, "") for header in DELIVERY_HEADERS])
    workbook.save(path)


def main() -> int:
    args = parse_args()
    source_root = Path(args.source).resolve()
    if not source_root.exists():
        print(f"[error] source does not exist: {source_root}")
        return 1

    output_root = (
        Path(args.output).resolve()
        if args.output
        else source_root.parent / f"{source_root.name}_{DEFAULT_OUTPUT_SUFFIX}"
    )
    report_path = output_root.parent / f"{output_root.name}_{REPORT_SUFFIX}.json"

    if output_root.exists():
        shutil.rmtree(output_root)
    output_root.mkdir(parents=True, exist_ok=True)

    submission_metadata = read_submission_metadata(source_root)
    report = {
        "source": str(source_root),
        "output": str(output_root),
        "fallback_submitter": args.submitter,
        "kept_sessions": [],
        "skipped_sessions": [],
    }
    delivery_rows: list[dict] = []

    for agent_name, agent_root in iter_agent_roots(source_root):
        grouped: dict[str, list[Path]] = defaultdict(list)
        for item in sorted(agent_root.iterdir()):
            if item.is_dir():
                grouped[item.name.split("__", 1)[0]].append(item)

        kept_for_agent = 0
        for session_id, candidates in sorted(grouped.items()):
            chosen = choose_best_candidate(candidates)
            target_dir = output_root / agent_name / session_id
            target_dir.parent.mkdir(parents=True, exist_ok=True)
            shutil.copytree(chosen.path, target_dir)

            normalize_session_json_files(target_dir, agent_name)
            normalize_label_file(target_dir)

            ok, errors, warnings, stats = validate_session(target_dir)
            if not ok:
                shutil.rmtree(target_dir)
                report["skipped_sessions"].append(
                    {
                        "agent": agent_name,
                        "session_id": session_id,
                        "source_dir": chosen.path.name,
                        "errors": errors,
                        "warnings": warnings,
                    }
                )
                continue

            kept_for_agent += 1
            metadata = submission_metadata.get(session_id, {})
            delivery_rows.append(
                {
                    HEADER_GROUP: agent_name,
                    HEADER_SUBMITTER: metadata.get("submitter") or args.submitter,
                    HEADER_SESSION_ID: session_id,
                    HEADER_SCENE: metadata.get("scene_name", ""),
                }
            )
            report["kept_sessions"].append(
                {
                    "agent": agent_name,
                    "session_id": session_id,
                    "source_dir": chosen.path.name,
                    "submitter": metadata.get("submitter") or args.submitter,
                    "warnings": warnings,
                    "assistant_turns": stats.get("assistant_turns", 0),
                }
            )

        if kept_for_agent == 0:
            empty_agent_dir = output_root / agent_name
            if empty_agent_dir.exists():
                shutil.rmtree(empty_agent_dir)

    delivery_rows.sort(key=lambda item: (str(item.get(HEADER_SESSION_ID, "")), str(item.get(HEADER_GROUP, ""))))
    write_delivery_xlsx(output_root / DELIVERY_XLSX_NAME, delivery_rows)

    report["summary"] = {
        "input_agent_dirs": [agent for agent, _ in iter_agent_roots(source_root)],
        "kept_count": len(report["kept_sessions"]),
        "skipped_count": len(report["skipped_sessions"]),
    }
    with report_path.open("w", encoding="utf-8") as handle:
        json.dump(report, handle, ensure_ascii=False, indent=2)

    print(f"[done] {output_root}")
    print(f"[report] {report_path}")
    print(f"[summary] kept={report['summary']['kept_count']} skipped={report['summary']['skipped_count']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
